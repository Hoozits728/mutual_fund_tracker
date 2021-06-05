import requests, bs4
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
#from selenium.webdriver import Firefox
#from selenium.webdriver.firefox.options import Options

#options = webdriver.ChromeOptions()
#options=Options()
#options.add_argument('--headless')
#options.add_argument('--disable-software-rasterizer')
#options.add_argument('--ignore-certificate-errors')
#options.add_argument('--ignore-ssl-errors')
#options.add_argument('--ignore-certificate-errors-spki-list')
#options.add_argument('--disable-gpu')
#options.add_argument('--window-size=800,600')
#options.add_argument('--allow-running-insecure-content')
#driver=webdriver.Chrome(r"C:\Users\light\Desktop\lightsquaresolutions\Mutual_Fund_Tracker\chromedriver.exe", chrome_options=options)
#driver=Firefox()
driver=webdriver.PhantomJS()
#wb=load_workbook(r'C:\Users\light\Desktop\lightsquaresolutions\Mutual_Fund_Tracker\Mutual_Fund_Tracker.xlsx')
wb=load_workbook(r'Mutual_Fund_Tracker.xlsx')
ws=wb.get_sheet_by_name('Mutual Funds')

for i in range(3, ws.max_row+1):
    if ws.cell(i,2).value is not None:
        #time.sleep(5)
        #print(ws.cell(i,2).value)
        res=requests.get('http://finance.yahoo.com/q?s=' + ws.cell(i,2).value + '&q1=1')
        soup=bs4.BeautifulSoup(res.text)
        try:
            mf_name=soup.select('#quote-header-info > div > div > div > h1')[0].text
        except:
            mf_name='None'
        try:
            webcat=soup.select('#quote-summary > div:nth-of-type(1) > table:nth-of-type(1) > tbody > tr:nth-of-type(4) > td:nth-of-type(2) > span')[0].text
        except:
            webcat='None'
        try:
            expenses=soup.select('#quote-summary > div:nth-of-type(1) > table:nth-of-type(1) > tbody > tr:nth-of-type(3) > td:nth-of-type(2) > span')[0].text
        except:
            expenses='None'
        try:
            total_assets = soup.select('#quote-summary > div:nth-of-type(2) > table > tbody > tr:nth-of-type(1) > td:nth-of-type(2) > span')[0].text
        except:
            total_assets='None'
        try:    
            ttmyield = soup.select('#quote-summary > div:nth-of-type(2) > table > tbody > tr:nth-of-type(3) > td:nth-of-type(2) > span')[0].text
        except:
            ttmyield='None'
        try:    
            webytd = soup.select('#quote-summary > div:nth-of-type(1) > table > tbody > tr:nth-of-type(2) > td:nth-of-type(2) > span')[0].text
        except:
            webytd='None'
        ws.cell(i,1).value=mf_name
        ws.cell(i,3).value=webcat
        ws.cell(i,6).value=expenses
        ws.cell(i,7).value=total_assets
        ws.cell(i,8).value=ttmyield
        ws.cell(i,9).value=webytd
        skip=False
        try:
            driver.get('http://finance.yahoo.com/quote/' + ws.cell(i,2).value + '/performance?ltr=1')
        except:
            skip=True
        stop=0
        start=time.time()
        while True and stop-start<10 and not skip:
            try:
                stop=time.time()
                oneyear=driver.find_element_by_css_selector('#Col1-0-Performance-Proxy > section > div:nth-child(2) > div > div:nth-child(5) > span:nth-child(2)').text  
                threeyear=driver.find_element_by_css_selector('#Col1-0-Performance-Proxy > section > div:nth-of-type(2) > div > div:nth-of-type(6) > span:nth-of-type(2)').text
                fiveyear=driver.find_element_by_css_selector('#Col1-0-Performance-Proxy > section > div:nth-child(2) > div > div:nth-child(7) > span.W\(20\%\).D\(b\).Fl\(start\).Ta\(e\)').text
                tenyear=driver.find_element_by_css_selector('#Col1-0-Performance-Proxy > section > div:nth-child(2) > div > div:nth-child(8) > span.W\(20\%\).D\(b\).Fl\(start\).Ta\(e\)').text
                ws.cell(i,10).value=oneyear
                ws.cell(i,11).value=threeyear
                ws.cell(i,12).value=fiveyear
                ws.cell(i,13).value=tenyear
                break
            except:
                pass
        skip=False
        try:
            driver.get('http://finance.yahoo.com/quote/' + ws.cell(i,2).value + '/profile?ltr=1')
        except:
            skip=True
        stop=0
        start=time.time()
        while True and stop-start<10 and not skip:
            try:
                stop=time.time()
                management=driver.find_element_by_css_selector('#Col1-0-Profile-Proxy > section > div > div > div:nth-of-type(4) > span').text.replace(",", "")
                mngmt_and_sd=driver.find_element_by_css_selector('#Col1-0-Profile-Proxy > section > div > div > div:nth-of-type(4) > span:nth-of-type(2) > span').text
                start_date=mngmt_and_sd
                ws.cell(i,4).value=start_date
                ws.cell(i,5).value=management
                break
            except:
                pass
        skip=False        
        try:
            driver.get('http://finance.yahoo.com/quote/' + ws.cell(i,2).value + '/risk?ltr=1')
        except:
            skip=True
        stop=0
        start=time.time()
        while True and stop-start<10 and not skip:
            try:
                stop=time.time()
                threeyearalpha=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(3) > div:nth-child(2) > span.W\(39\%\).Fl\(start\)').text
                threeyearbeta=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(4) > div:nth-child(2) > span.W\(39\%\).Fl\(start\)').text
                threeyearrsquared=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(6) > div:nth-child(2) > span.W\(39\%\).Fl\(start\)').text
                threeyearsd=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(7) > div:nth-child(2) > span.W\(39\%\).Fl\(start\)').text
                threeyearsharpe=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(8) > div:nth-child(2) > span.W\(39\%\).Fl\(start\)').text
                ws.cell(i,19).value=threeyearalpha
                ws.cell(i,18).value=threeyearbeta
                ws.cell(i,17).value=threeyearrsquared
                ws.cell(i,14).value=threeyearsd
                ws.cell(i,15).value=threeyearsharpe 
               
                fiveyearalpha=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(3) > div:nth-child(3) > span.W\(39\%\).Fl\(start\)').text
                fiveyearbeta=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(4) > div:nth-child(3) > span.W\(39\%\).Fl\(start\)').text
                fiveyearrsquared=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(6) > div:nth-child(3) > span.W\(39\%\).Fl\(start\)').text
                fiveyearsd=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(7) > div:nth-child(3) > span.W\(39\%\).Fl\(start\)').text
                fiveyearsharpe=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(8) > div:nth-child(3) > span.W\(39\%\).Fl\(start\)').text
                ws.cell(i,25).value=fiveyearalpha
                ws.cell(i,24).value=fiveyearbeta
                ws.cell(i,23).value=fiveyearrsquared
                ws.cell(i,20).value=fiveyearsd
                ws.cell(i,21).value=fiveyearsharpe        

                tenyearalpha=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(3) > div:nth-child(4) > span.W\(39\%\).Fl\(start\)').text
                tenyearbeta=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(4) > div:nth-child(4) > span.W\(39\%\).Fl\(start\)').text
                tenyearrsquared=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(6) > div:nth-child(4) > span.W\(39\%\).Fl\(start\)').text
                tenyearsd=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(7) > div:nth-child(4) > span.W\(39\%\).Fl\(start\)').text
                tenyearsharpe=driver.find_element_by_css_selector('#Col1-0-Risk-Proxy > section > div.Mb\(25px\).Ovx\(a\) > div > div:nth-child(8) > div:nth-child(4) > span.W\(39\%\).Fl\(start\)').text
                ws.cell(i,31).value=tenyearalpha
                ws.cell(i,30).value=tenyearbeta
                ws.cell(i,29).value=tenyearrsquared
                ws.cell(i,26).value=tenyearsd
                ws.cell(i,27).value=tenyearsharpe   
                break
            except:
                pass
        
        #time.sleep(5)
        skip=False
        try:
            driver.get('http://performance.morningstar.com/fund/tax-analysis.action?t=' + ws.cell(i,2).value + '&region=usa&culture=en-US')
        except:
            skip=True
        #time.sleep(5)
        stop=0
        start=time.time()
        while True and stop-start<10 and not skip:    
            try:
                #time.sleep(3)
                stop=time.time()
                #print(stop-start)
                ytdtax=driver.find_element_by_css_selector('#div_tax_analysis > table > tbody > tr:nth-child(8) > td:nth-child(5)').text
                oneyeartax=driver.find_element_by_css_selector('#div_tax_analysis > table > tbody > tr:nth-child(8) > td:nth-child(6)').text
                threeyeartax=driver.find_element_by_css_selector('#div_tax_analysis > table > tbody > tr:nth-child(8) > td:nth-child(7)').text
                fiveyeartax = driver.find_element_by_css_selector('#div_tax_analysis > table > tbody > tr:nth-child(8) > td:nth-child(8)').text
                tenyeartax = driver.find_element_by_css_selector('#div_tax_analysis > table > tbody > tr:nth-child(8) > td:nth-child(9)').text
                tenyear = driver.find_element_by_css_selector('#div_tax_analysis > table > tbody > tr:nth-child(4) > td:nth-child(9)').text
                ws.cell(i,32).value=ytdtax
                ws.cell(i,33).value=oneyeartax
                ws.cell(i,34).value=threeyeartax
                ws.cell(i,35).value=fiveyeartax
                ws.cell(i,36).value=tenyeartax
                #ws.cell(i,13).value=tenyear  
                break
            except:  
                pass
                
        #time.sleep(5)    
        skip=False
        try:
            driver.get('http://performance.morningstar.com/fund/ratings-risk.action?t=' + ws.cell(i,2).value + '&region=usa&culture=en-US')
        except:
            skip=True
        #while True:
        if not skip:
            try:
                element = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#year3vm")))
            #driver.find_element_by_css_selector('#year3vm').click()
                element.click()
            #        break
                

                element_present=EC.presence_of_element_located((By.ID, 'div_volatility'))
                WebDriverWait(driver, 10).until(element_present)

                while True:
                    try:
                        threeyearsortino=driver.find_element_by_css_selector('#div_volatility > table > tbody > tr:nth-child(2) > td:nth-child(5)').text
                        ws.cell(i,16).value=threeyearsortino
                        break
                    except:
                        pass

                #time.sleep(3)
                element = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#year5vm")))
                element.click()
                #driver.find_element_by_css_selector('#year5vm').click()

                element_present=EC.presence_of_element_located((By.ID, 'div_volatility'))
                WebDriverWait(driver, 10).until(element_present)

                while True:
                    try:
                        fiveyearsortino=driver.find_element_by_css_selector('#div_volatility > table > tbody > tr:nth-child(2) > td:nth-child(5)').text
                        ws.cell(i,22).value=fiveyearsortino
                        break
                    except:
                        pass
                #time.sleep(3)
                element = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#year10vm")))
                element.click()
                #driver.find_element_by_css_selector('#year10vm').click()

                element_present=EC.presence_of_element_located((By.ID, 'div_volatility'))
                WebDriverWait(driver, 10).until(element_present)

                while True:
                    try:
                        tenyearsortino=driver.find_element_by_css_selector('#div_volatility > table > tbody > tr:nth-child(2) > td:nth-child(5)').text
                        ws.cell(i,28).value=tenyearsortino
                        break
                    except:
                        pass
            
            except:
                pass

#wb.save(r'C:\Users\light\Desktop\lightsquaresolutions\Mutual_Fund_Tracker\Mutual_Fund_Tracker.xlsx')
wb.save(r'Mutual_Fund_Tracker.xlsx')
driver.quit()   